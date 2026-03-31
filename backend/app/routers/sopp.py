"""
S&OP Router — lê as tabelas BI do banco BISOBEL (read-only).
Usa pymssql: parâmetros com %s, literal % precisa de %% em queries parametrizadas.
"""
import io
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.database_bi import get_bi_conn, rows_to_dicts
from app.database import get_db
from app.deps import get_current_user
from app import models

router = APIRouter(prefix="/sopp", tags=["sopp"])

BRAND_RED = "D92214"

# ---------------------------------------------------------------------------
# Linha / Formato — mapeamento S&OP para carteira de pedidos
# ---------------------------------------------------------------------------

_LINHA_CASE = """
    CASE
      WHEN LEFT(DESC_PRD,14) = 'AGUA SANITARIA'  THEN 'Agua Sanitaria'
      WHEN LEFT(DESC_PRD,9)  = 'ALVEJANTE'        THEN 'Alvejante'
      WHEN LEFT(DESC_PRD,4)  = 'AMAC'             THEN 'Amaciante'
      WHEN LEFT(DESC_PRD,5)  = 'CLORO'            THEN 'Cloro'
      WHEN LEFT(DESC_PRD,6)  = 'DESINF'           THEN 'Desinfetante'
      WHEN LEFT(DESC_PRD,9)  = 'LAVA LOUC'        THEN 'Lava Loucas'
      WHEN LEFT(DESC_PRD,10) = 'LAVA ROUPA'       THEN 'Lava Roupas'
      WHEN LEFT(DESC_PRD,4)  = 'LIMP'
        OR LEFT(DESC_PRD,15) = 'DESENGORDURANTE'  THEN 'Limpador'
      WHEN LEFT(DESC_PRD,9)  = 'REMOVEDOR'        THEN 'Removedor'
      ELSE 'Outros'
    END
"""

# Mesma lógica para MOV_PRODUCAO onde a coluna se chama PRODUTO (não DESC_PRD)
_LINHA_CASE_PRODUTO = """
    CASE
      WHEN LEFT(PRODUTO,14) = 'AGUA SANITARIA'  THEN 'Agua Sanitaria'
      WHEN LEFT(PRODUTO,9)  = 'ALVEJANTE'        THEN 'Alvejante'
      WHEN LEFT(PRODUTO,4)  = 'AMAC'             THEN 'Amaciante'
      WHEN LEFT(PRODUTO,5)  = 'CLORO'            THEN 'Cloro'
      WHEN LEFT(PRODUTO,6)  = 'DESINF'           THEN 'Desinfetante'
      WHEN LEFT(PRODUTO,9)  = 'LAVA LOUC'        THEN 'Lava Loucas'
      WHEN LEFT(PRODUTO,10) = 'LAVA ROUPA'       THEN 'Lava Roupas'
      WHEN LEFT(PRODUTO,4)  = 'LIMP'
        OR LEFT(PRODUTO,15) = 'DESENGORDURANTE'  THEN 'Limpador'
      WHEN LEFT(PRODUTO,9)  = 'REMOVEDOR'        THEN 'Removedor'
      ELSE 'Outros'
    END
"""

_FORMATO_CASE = """
    CASE
      WHEN CHARINDEX('X500',COD_PRD)  > 0                                      THEN '500ml'
      WHEN CHARINDEX('X01L',COD_PRD)  > 0                                      THEN '1L'
      WHEN CHARINDEX('X1,5',COD_PRD)  > 0 OR CHARINDEX('X1.5',COD_PRD) > 0    THEN '1,5L'
      WHEN CHARINDEX('X02L',COD_PRD)  > 0 OR CHARINDEX('X2L', COD_PRD) > 0    THEN '2L'
      WHEN CHARINDEX('X03L',COD_PRD)  > 0                                      THEN '3L'
      WHEN CHARINDEX('X05L',COD_PRD)  > 0                                      THEN '5L'
      ELSE 'Outros'
    END
"""

# Traduz valor de filtro para condição SQL literal (sem params — valores são de lista controlada)
_LINHA_COND = {
    'Agua Sanitaria':  "LEFT(DESC_PRD,14)='AGUA SANITARIA'",
    'Alvejante':       "LEFT(DESC_PRD,9)='ALVEJANTE'",
    'Amaciante':       "LEFT(DESC_PRD,4)='AMAC'",
    'Cloro':           "LEFT(DESC_PRD,5)='CLORO'",
    'Desinfetante':    "LEFT(DESC_PRD,6)='DESINF'",
    'Lava Loucas':     "LEFT(DESC_PRD,9)='LAVA LOUC'",
    'Lava Roupas':     "LEFT(DESC_PRD,10)='LAVA ROUPA'",
    'Limpador':        "(LEFT(DESC_PRD,4)='LIMP' OR LEFT(DESC_PRD,15)='DESENGORDURANTE')",
    'Removedor':       "LEFT(DESC_PRD,9)='REMOVEDOR'",
}

_FORMATO_COND = {
    '500ml': "CHARINDEX('X500',COD_PRD)>0",
    '1L':    "CHARINDEX('X01L',COD_PRD)>0",
    '1,5L':  "(CHARINDEX('X1,5',COD_PRD)>0 OR CHARINDEX('X1.5',COD_PRD)>0)",
    '2L':    "(CHARINDEX('X02L',COD_PRD)>0 OR CHARINDEX('X2L',COD_PRD)>0)",
    '3L':    "CHARINDEX('X03L',COD_PRD)>0",
    '5L':    "CHARINDEX('X05L',COD_PRD)>0",
}

LINHAS_PV   = list(_LINHA_COND.keys())
FORMATOS_PV = list(_FORMATO_COND.keys())

# Mesma lógica para FATURAMENTO onde a coluna se chama DESCRICAO
_LINHA_CASE_DESCRICAO = """
    CASE
      WHEN LEFT(DESCRICAO,14) = 'AGUA SANITARIA'  THEN 'Agua Sanitaria'
      WHEN LEFT(DESCRICAO,9)  = 'ALVEJANTE'        THEN 'Alvejante'
      WHEN LEFT(DESCRICAO,4)  = 'AMAC'             THEN 'Amaciante'
      WHEN LEFT(DESCRICAO,5)  = 'CLORO'            THEN 'Cloro'
      WHEN LEFT(DESCRICAO,6)  = 'DESINF'           THEN 'Desinfetante'
      WHEN LEFT(DESCRICAO,9)  = 'LAVA LOUC'        THEN 'Lava Loucas'
      WHEN LEFT(DESCRICAO,10) = 'LAVA ROUPA'       THEN 'Lava Roupas'
      WHEN LEFT(DESCRICAO,4)  = 'LIMP'
        OR LEFT(DESCRICAO,15) = 'DESENGORDURANTE'  THEN 'Limpador'
      WHEN LEFT(DESCRICAO,9)  = 'REMOVEDOR'        THEN 'Removedor'
      ELSE 'Outros'
    END
"""

_LINHA_COND_DESCRICAO = {
    'Agua Sanitaria':  "LEFT(DESCRICAO,14)='AGUA SANITARIA'",
    'Alvejante':       "LEFT(DESCRICAO,9)='ALVEJANTE'",
    'Amaciante':       "LEFT(DESCRICAO,4)='AMAC'",
    'Cloro':           "LEFT(DESCRICAO,5)='CLORO'",
    'Desinfetante':    "LEFT(DESCRICAO,6)='DESINF'",
    'Lava Loucas':     "LEFT(DESCRICAO,9)='LAVA LOUC'",
    'Lava Roupas':     "LEFT(DESCRICAO,10)='LAVA ROUPA'",
    'Limpador':        "(LEFT(DESCRICAO,4)='LIMP' OR LEFT(DESCRICAO,15)='DESENGORDURANTE')",
    'Removedor':       "LEFT(DESCRICAO,9)='REMOVEDOR'",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _exec(cur, sql: str, params: list | tuple = ()):
    """Executa query; só passa params ao cursor quando necessário (evita
    processamento de % pelo pymssql em queries sem parâmetros)."""
    if params:
        cur.execute(sql, list(params))
    else:
        cur.execute(sql)


def _scalar(conn, sql: str, params=()) -> float:
    cur = conn.cursor()
    _exec(cur, sql, params)
    val = cur.fetchone()[0]
    return float(val) if val is not None else 0.0


def _build_xlsx(sheet_title: str, headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    hdr_fill = PatternFill("solid", fgColor=BRAND_RED)
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, bottom=thin)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22

    for ri, row_data in enumerate(rows, 2):
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=val).border = border

    for ci, h in enumerate(headers, 1):
        col_letter = get_column_letter(ci)
        max_len = len(h)
        for ri in range(2, min(len(rows) + 2, 102)):
            cell_val = str(ws.cell(row=ri, column=ci).value or "")
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _xlsx_response(data: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------------------------

@router.get("/dashboard")
def get_dashboard(
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    kpis = {
        "qtde_carteira":    _scalar(conn, "SELECT ISNULL(SUM(QTDE),0) FROM PV_ABERTOS"),
        "valor_carteira":   _scalar(conn, "SELECT ISNULL(SUM(VL_BRUTO),0) FROM PV_ABERTOS"),
        "estoque_disponivel": _scalar(
            conn,
            "SELECT ISNULL(SUM(QTDE_DISP),0) FROM ESTOQUES_PA WHERE TB_EMPRESA='010'"
        ),
        "fat_mes_atual": _scalar(conn, """
            SELECT ISNULL(SUM(VL_LIQUIDO),0) FROM FATURAMENTO
            WHERE EMISSAO >= DATEADD(month,DATEDIFF(month,0,GETDATE()),0)
              AND TP = 'VS'
        """),
        "qtde_fat_mes": _scalar(conn, """
            SELECT ISNULL(SUM(QUANTIDADE),0) FROM FATURAMENTO
            WHERE EMISSAO >= DATEADD(month,DATEDIFF(month,0,GETDATE()),0)
              AND TP = 'VS'
        """),
        "fat_devol_mes": _scalar(conn, """
            SELECT ISNULL(SUM(ABS(VL_LIQUIDO)),0) FROM FATURAMENTO
            WHERE EMISSAO >= DATEADD(month,DATEDIFF(month,0,GETDATE()),0)
              AND TP = 'DS'
        """),
        "qtde_devol_mes": _scalar(conn, """
            SELECT ISNULL(SUM(ABS(QUANTIDADE)),0) FROM FATURAMENTO
            WHERE EMISSAO >= DATEADD(month,DATEDIFF(month,0,GETDATE()),0)
              AND TP = 'DS'
        """),
        "fat_bonif_mes": _scalar(conn, """
            SELECT ISNULL(SUM(VL_LIQUIDO),0) FROM FATURAMENTO
            WHERE EMISSAO >= DATEADD(month,DATEDIFF(month,0,GETDATE()),0)
              AND TP = 'FS'
        """),
        "qtde_bonif_mes": _scalar(conn, """
            SELECT ISNULL(SUM(QUANTIDADE),0) FROM FATURAMENTO
            WHERE EMISSAO >= DATEADD(month,DATEDIFF(month,0,GETDATE()),0)
              AND TP = 'FS'
        """),
        "prod_mes_atual": _scalar(conn, """
            SELECT ISNULL(SUM(QTDE),0) FROM MOV_PRODUCAO
            WHERE TIPO='PA'
              AND EMISSAO >= DATEADD(month,DATEDIFF(month,0,GETDATE()),0)
        """),
    }

    cur = conn.cursor()

    # Faturamento mensal: líquido (TP=VS) + devoluções (TP=DS, ABS)
    _exec(cur, """
        SELECT FORMAT(EMISSAO,'yyyy-MM') as mes,
               ISNULL(SUM(CASE WHEN TP = 'VS' THEN VL_LIQUIDO ELSE 0 END), 0) as liquido,
               ISNULL(SUM(CASE WHEN TP = 'DS' THEN ABS(VL_LIQUIDO) ELSE 0 END), 0) as devol
        FROM FATURAMENTO
        WHERE EMISSAO >= DATEADD(month,-5,DATEADD(month,DATEDIFF(month,0,GETDATE()),0))
        GROUP BY FORMAT(EMISSAO,'yyyy-MM')
        ORDER BY mes
    """)
    fat_mensal = [
        {"mes": r[0], "liquido": float(r[1] or 0), "devol": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    # Carteira por linha S&OP (usando DESC_PRD → _LINHA_CASE)
    _exec(cur, f"""
        SELECT linha,
               ISNULL(SUM(valor),0) as valor,
               ISNULL(SUM(qtde),0)  as qtde
        FROM (
            SELECT {_LINHA_CASE} as linha,
                   ISNULL(VL_BRUTO,0) as valor,
                   ISNULL(QTDE,0)     as qtde
            FROM PV_ABERTOS
        ) t
        GROUP BY linha
        ORDER BY valor DESC
    """)
    carteira_por_linha = [
        {"linha": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    _exec(cur, """
        SELECT FORMAT(EMISSAO,'yyyy-MM-dd') as data, ISNULL(SUM(QTDE),0) as qtde
        FROM MOV_PRODUCAO
        WHERE TIPO='PA' AND EMISSAO >= DATEADD(day,-30,GETDATE())
        GROUP BY FORMAT(EMISSAO,'yyyy-MM-dd')
        ORDER BY data
    """)
    producao_diaria = [{"data": r[0], "qtde": float(r[1] or 0)} for r in cur.fetchall()]

    _exec(cur, """
        SELECT TOP 10 COD_PRD, DESC_PRD,
               ISNULL(SUM(QTDE_DISP),0) as disp,
               ISNULL(SUM(QTDE_RESERVA),0) as reserva
        FROM ESTOQUES_PA
        WHERE TB_EMPRESA='010' AND QTDE_DISP > 0
        GROUP BY COD_PRD, DESC_PRD
        ORDER BY disp DESC
    """)
    estoque_top10 = [
        {"cod_prd": r[0], "desc_prd": r[1],
         "disp": float(r[2] or 0), "reserva": float(r[3] or 0)}
        for r in cur.fetchall()
    ]

    return {
        "kpis": kpis,
        "fat_mensal": fat_mensal,
        "carteira_por_linha": carteira_por_linha,
        "producao_diaria": producao_diaria,
        "estoque_top10": estoque_top10,
    }


# ---------------------------------------------------------------------------
# CARTEIRA DE PEDIDOS (PV_ABERTOS)
# ---------------------------------------------------------------------------

def _pedidos_where(linha, formato, search, date_from, date_to):
    conds, params = [], []
    # linha e formato: valores de lista controlada — usados como literais SQL
    if linha and linha in _LINHA_COND:
        conds.append(_LINHA_COND[linha])
    if formato and formato in _FORMATO_COND:
        conds.append(_FORMATO_COND[formato])
    if search:
        conds.append(
            "(COD_PRD LIKE %s OR DESC_PRD LIKE %s OR CLIENTE LIKE %s)"
        )
        p = f"%{search}%"
        params += [p, p, p]
    if date_from:
        conds.append("DT_ENTREGA >= %s")
        params.append(date_from)
    if date_to:
        conds.append("DT_ENTREGA <= %s")
        params.append(date_to)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    return where, params


@router.get("/pedidos")
def list_pedidos(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    linha: str | None = Query(None),
    formato: str | None = Query(None),
    search: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, params = _pedidos_where(linha, formato, search, date_from, date_to)
    cur = conn.cursor()
    _exec(cur, f"SELECT COUNT(*) FROM PV_ABERTOS {where}", params)
    total = cur.fetchone()[0]

    offset = (page - 1) * page_size
    _exec(cur, f"""
        SELECT SK_PEDIDO, NUM_PV, COD_PRD, DESC_PRD, FAM,
               QTDE, VL_BRUTO, CLIENTE, VEND, NOME_VEND, SUP,
               CONVERT(varchar,EMISSAO,23) as EMISSAO,
               CONVERT(varchar,DT_AGENDA,23) as DT_AGENDA,
               CONVERT(varchar,DT_ENTREGA,23) as DT_ENTREGA
        FROM PV_ABERTOS {where}
        ORDER BY DT_ENTREGA ASC, SK_PEDIDO
        OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY
    """, params)
    items = rows_to_dicts(cur)

    return {
        "items": items, "total": total, "page": page,
        "page_size": page_size,
        "total_pages": max(1, math.ceil(total / page_size)),
    }


@router.get("/pedidos/charts")
def pedidos_charts(
    linha: str | None = Query(None),
    formato: str | None = Query(None),
    search: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, params = _pedidos_where(linha, formato, search, date_from, date_to)
    cur = conn.cursor()

    # KPIs agregados
    _exec(cur, f"""
        SELECT COUNT(*), ISNULL(SUM(QTDE),0), ISNULL(SUM(VL_BRUTO),0)
        FROM PV_ABERTOS {where}
    """, params)
    r = cur.fetchone()
    total_pedidos = int(r[0] or 0)
    total_qtde = float(r[1] or 0)
    total_valor = float(r[2] or 0)
    preco_medio = total_valor / total_qtde if total_qtde > 0 else 0

    # Por linha de produto — volume S&OP
    _exec(cur, f"""
        SELECT ({_LINHA_CASE}) as linha,
               ISNULL(SUM(QTDE),0) as qtde,
               ISNULL(SUM(VL_BRUTO),0) as valor
        FROM PV_ABERTOS {where}
        GROUP BY ({_LINHA_CASE})
        ORDER BY qtde DESC
    """, params)
    por_linha = [
        {"linha": r[0], "qtde": float(r[1] or 0), "valor": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    # Por formato/volume — volume S&OP
    _exec(cur, f"""
        SELECT ({_FORMATO_CASE}) as formato,
               ISNULL(SUM(QTDE),0) as qtde,
               ISNULL(SUM(VL_BRUTO),0) as valor
        FROM PV_ABERTOS {where}
        GROUP BY ({_FORMATO_CASE})
        ORDER BY qtde DESC
    """, params)
    por_formato = [
        {"formato": r[0], "qtde": float(r[1] or 0), "valor": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    return {
        "kpis": {
            "total_pedidos": total_pedidos,
            "total_qtde": total_qtde,
            "total_valor": total_valor,
            "preco_medio": preco_medio,
        },
        "por_linha": por_linha,
        "por_formato": por_formato,
    }


@router.get("/pedidos/export")
def export_pedidos(
    linha: str | None = Query(None),
    formato: str | None = Query(None),
    search: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, params = _pedidos_where(linha, formato, search, date_from, date_to)
    cur = conn.cursor()
    _exec(cur, f"""
        SELECT SK_PEDIDO, NUM_PV, COD_PRD, DESC_PRD, FAM,
               QTDE, VL_BRUTO, CLIENTE, NOME_VEND, SUP,
               CONVERT(varchar,EMISSAO,103), CONVERT(varchar,DT_ENTREGA,103)
        FROM PV_ABERTOS {where}
        ORDER BY DT_ENTREGA ASC
    """, params)
    headers = ["PV", "Nº PV", "Produto", "Descrição", "Família",
               "Qtde", "Valor Bruto", "Cliente", "Vendedor", "Supervisor",
               "Emissão", "Entrega Prev."]
    xlsx = _build_xlsx("Carteira de Pedidos", headers, [list(r) for r in cur.fetchall()])
    return _xlsx_response(xlsx, "carteira_pedidos.xlsx")


# ---------------------------------------------------------------------------
# PRODUÇÃO (MOV_PRODUCAO)
# ---------------------------------------------------------------------------

_FILIAL_LABEL_PROD = {
    "01":  "Jaçanã (01)",
    "010": "Jaçanã (01)",
    "06":  "Atibaia (06)",
    "060": "Atibaia (06)",
}


def _producao_where(tipo, cod_prd, date_from, date_to, operacao=None, filial=None):
    conds, params = [], []
    if date_from == "CURRENT_MONTH":
        conds.append("EMISSAO >= DATEADD(month,DATEDIFF(month,0,GETDATE()),0)")
    elif date_from:
        conds.append("EMISSAO >= %s"); params.append(date_from)
    if date_from != "CURRENT_MONTH" and date_to:
        conds.append("EMISSAO <= %s"); params.append(date_to)
    if tipo:
        conds.append("TIPO = %s"); params.append(tipo)
    if operacao:
        conds.append("OPERACAO = %s"); params.append(operacao)
    if cod_prd:
        conds.append("(COD_PRD LIKE %s OR PRODUTO LIKE %s)")
        p = f"%{cod_prd}%"; params += [p, p]
    if filial:
        conds.append("LTRIM(RTRIM(D3_FILIAL)) = %s"); params.append(filial)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    return where, params


@router.get("/producao")
def list_producao(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    tipo: str | None = Query(None),
    cod_prd: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    filial: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    if not tipo:
        tipo = "PA"
    if not date_from and not date_to:
        date_from = "CURRENT_MONTH"
    where, params = _producao_where(tipo, cod_prd, date_from, date_to, operacao="PR0", filial=filial)
    cur = conn.cursor()
    _exec(cur, f"SELECT COUNT(*) FROM MOV_PRODUCAO {where}", params)
    total = cur.fetchone()[0]

    offset = (page - 1) * page_size
    _exec(cur, f"""
        SELECT SK_MOVTO, OP, LOTE, COD_PRD, PRODUTO, TIPO, OPERACAO,
               QTDE, ARMAZEM,
               CONVERT(varchar,EMISSAO,23) as EMISSAO,
               CONVERT(varchar,HRA,114) as HRA,
               D3_FILIAL as TB_FILIAL
        FROM MOV_PRODUCAO {where}
        ORDER BY EMISSAO DESC, SK_MOVTO DESC
        OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY
    """, params)
    items = rows_to_dicts(cur)
    return {
        "items": items, "total": total, "page": page,
        "page_size": page_size,
        "total_pages": max(1, math.ceil(total / page_size)),
    }


@router.get("/producao/charts")
def producao_charts(
    tipo: str | None = Query(None),
    cod_prd: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    filial: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    if not tipo:
        tipo = "PA"
    if not date_from and not date_to:
        date_from = "CURRENT_MONTH"
    # where com filial (para KPIs / gráficos da seleção)
    where, params = _producao_where(tipo, cod_prd, date_from, date_to, operacao="PR0", filial=filial)
    # where base sem filial (para breakdown por filial — sempre mostra as duas)
    where_base, params_base = _producao_where(tipo, cod_prd, date_from, date_to, operacao="PR0")
    cur = conn.cursor()

    # KPIs agregados
    _exec(cur, f"""
        SELECT ISNULL(SUM(QTDE),0),
               COUNT(DISTINCT FORMAT(EMISSAO,'yyyy-MM-dd')),
               COUNT(*)
        FROM MOV_PRODUCAO {where}
    """, params)
    r = cur.fetchone()
    total_qtde = float(r[0] or 0)
    dias_produtivos = int(r[1] or 0)
    total_movtos = int(r[2] or 0)
    media_dia = total_qtde / dias_produtivos if dias_produtivos > 0 else 0

    # Produção diária PA
    _exec(cur, f"""
        SELECT FORMAT(EMISSAO,'yyyy-MM-dd') as data, ISNULL(SUM(QTDE),0) as qtde
        FROM MOV_PRODUCAO {where}
        GROUP BY FORMAT(EMISSAO,'yyyy-MM-dd')
        ORDER BY data
    """, params)
    por_dia = [{"data": r[0], "qtde": float(r[1] or 0)} for r in cur.fetchall()]

    # Top produto (para KPI card)
    _exec(cur, f"""
        SELECT TOP 1 COD_PRD, PRODUTO, ISNULL(SUM(QTDE),0) as qtde
        FROM MOV_PRODUCAO {where}
        GROUP BY COD_PRD, PRODUTO ORDER BY qtde DESC
    """, params)
    r1 = cur.fetchone()
    top1 = {"cod_prd": r1[0], "produto": r1[1], "qtde": float(r1[2] or 0)} if r1 else None

    # Por linha (família S&OP) via subquery para evitar ambiguidade no GROUP BY
    _exec(cur, f"""
        SELECT linha, ISNULL(SUM(qtde),0) as qtde
        FROM (
            SELECT {_LINHA_CASE_PRODUTO} as linha, QTDE as qtde
            FROM MOV_PRODUCAO {where}
        ) t
        GROUP BY linha
        ORDER BY qtde DESC
    """, params)
    por_linha = [
        {"linha": r[0], "qtde": float(r[1] or 0)}
        for r in cur.fetchall()
    ]

    # Breakdown por filial — sempre retorna 010 e 060, mesmo sem dados
    # Converte "WHERE cond1 AND cond2" → "AND cond1 AND cond2" para usar no ON do LEFT JOIN
    join_cond = where_base.replace("WHERE ", "AND ", 1) if where_base.startswith("WHERE ") else ""
    _exec(cur, f"""
        SELECT f.filial,
               ISNULL(SUM(m.QTDE), 0)                          as total_qtde,
               COUNT(DISTINCT CASE WHEN m.EMISSAO IS NOT NULL
                    THEN FORMAT(m.EMISSAO,'yyyy-MM-dd') END)    as dias_produtivos,
               COUNT(m.SK_MOVTO)                                as total_movtos
        FROM (VALUES ('01'),('06')) f(filial)
        LEFT JOIN MOV_PRODUCAO m
               ON LTRIM(RTRIM(m.D3_FILIAL)) = f.filial
              {join_cond}
        GROUP BY f.filial
        ORDER BY f.filial
    """, params_base)
    por_filial = []
    for row in cur.fetchall():
        tq = float(row[1] or 0)
        dp = int(row[2] or 0)
        por_filial.append({
            "filial":          row[0],
            "label":           _FILIAL_LABEL_PROD.get(row[0], f"Filial {row[0]}"),
            "total_qtde":      tq,
            "dias_produtivos": dp,
            "total_movtos":    int(row[3] or 0),
            "media_dia":       tq / dp if dp > 0 else 0,
        })

    return {
        "kpis": {
            "total_qtde": total_qtde,
            "dias_produtivos": dias_produtivos,
            "media_dia": media_dia,
            "total_movtos": total_movtos,
            "top_produto": top1,
        },
        "por_dia": por_dia,
        "por_linha": por_linha,
        "por_filial": por_filial,
    }


@router.get("/producao/export")
def export_producao(
    tipo: str | None = Query(None),
    cod_prd: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    filial: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    if not date_from and not date_to:
        date_from = "CURRENT_MONTH"
    where, params = _producao_where(tipo, cod_prd, date_from, date_to, operacao="PR0", filial=filial)
    cur = conn.cursor()
    _exec(cur, f"""
        SELECT TOP 10000
               CONVERT(varchar,EMISSAO,103), OP, LOTE,
               COD_PRD, PRODUTO, TIPO, OPERACAO, QTDE, ARMAZEM, D3_FILIAL
        FROM MOV_PRODUCAO {where}
        ORDER BY EMISSAO DESC
    """, params)
    headers = ["Data", "OP", "Lote", "Produto", "Descrição",
               "Tipo", "Operação", "Qtde", "Armazém", "Filial"]
    xlsx = _build_xlsx("Produção", headers, [list(r) for r in cur.fetchall()])
    return _xlsx_response(xlsx, "producao.xlsx")


# ---------------------------------------------------------------------------
# FORECAST (SKU_FORECAST)
# A coluna [%_ABC] tem % que conflita com pymssql em queries parametrizadas.
# Solução: _exec sem params passa SQL sem processamento de %; com params usa %%_ABC.
# ---------------------------------------------------------------------------

def _forecast_perc_col(has_params: bool) -> str:
    """Retorna o nome da coluna %_ABC seguro para pymssql."""
    return "[%%_ABC]" if has_params else "[%_ABC]"


@router.get("/forecast")
def list_forecast(
    linha: str | None = Query(None),
    search: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    conds, params = [], []
    if linha and linha in _LINHA_COND:
        conds.append(_LINHA_COND[linha])
    if search:
        conds.append("(COD_PRD LIKE %s OR DESC_PRD LIKE %s)")
        p = f"%{search}%"; params += [p, p]
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    perc = _forecast_perc_col(bool(params))

    cur = conn.cursor()
    sql = f"""
        SELECT SK_FORECAST, COD_PRD, LINHA, DESC_PRD, GRP_PRD, COD_FAM, FAM,
               MARCA, TIPO, {perc} as PERC_ABC,
               JAN, FEV, MAR, ABR, MAI, JUN, JUL, AGO, [SET] as SET_M,
               OUT, NOV, DEZ, TOTAL
        FROM SKU_FORECAST {where}
        ORDER BY FAM, COD_PRD
    """
    _exec(cur, sql, params)
    items = rows_to_dicts(cur)

    # Faturamento real ano corrente por produto
    _exec(cur, """
        SELECT COD_PRD,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=1  THEN QUANTIDADE END),0) as jan,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=2  THEN QUANTIDADE END),0) as fev,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=3  THEN QUANTIDADE END),0) as mar,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=4  THEN QUANTIDADE END),0) as abr,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=5  THEN QUANTIDADE END),0) as mai,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=6  THEN QUANTIDADE END),0) as jun,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=7  THEN QUANTIDADE END),0) as jul,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=8  THEN QUANTIDADE END),0) as ago,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=9  THEN QUANTIDADE END),0) as set_m,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=10 THEN QUANTIDADE END),0) as out_m,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=11 THEN QUANTIDADE END),0) as nov,
               ISNULL(SUM(CASE WHEN MONTH(EMISSAO)=12 THEN QUANTIDADE END),0) as dez,
               ISNULL(SUM(QUANTIDADE),0) as total
        FROM FATURAMENTO
        WHERE YEAR(EMISSAO) = YEAR(GETDATE())
          AND TP NOT IN ('DEV','1DEV') AND TIPO NOT LIKE '%DEV%'
        GROUP BY COD_PRD
    """)
    fat_real = {r["cod_prd"]: r for r in rows_to_dicts(cur)}

    return {"items": items, "fat_real": fat_real}


@router.get("/forecast/summary")
def forecast_summary(
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    cur = conn.cursor()

    # Forecast agrupado por linha S&OP (DESC_PRD)
    _exec(cur, f"""
        SELECT linha, ISNULL(SUM(fc),0) as forecast,
               ISNULL(SUM(jan),0) as jan, ISNULL(SUM(fev),0) as fev,
               ISNULL(SUM(mar),0) as mar, ISNULL(SUM(abr),0) as abr,
               ISNULL(SUM(mai),0) as mai, ISNULL(SUM(jun),0) as jun,
               ISNULL(SUM(jul),0) as jul, ISNULL(SUM(ago),0) as ago,
               ISNULL(SUM(set_m),0) as set_m, ISNULL(SUM(out_m),0) as out_m,
               ISNULL(SUM(nov),0) as nov, ISNULL(SUM(dez),0) as dez
        FROM (
            SELECT {_LINHA_CASE} as linha,
                   ISNULL(JAN+FEV+MAR+ABR+MAI+JUN+JUL+AGO+[SET]+OUT+NOV+DEZ,0) as fc,
                   ISNULL(JAN,0) as jan, ISNULL(FEV,0) as fev,
                   ISNULL(MAR,0) as mar, ISNULL(ABR,0) as abr,
                   ISNULL(MAI,0) as mai, ISNULL(JUN,0) as jun,
                   ISNULL(JUL,0) as jul, ISNULL(AGO,0) as ago,
                   ISNULL([SET],0) as set_m, ISNULL(OUT,0) as out_m,
                   ISNULL(NOV,0) as nov, ISNULL(DEZ,0) as dez
            FROM SKU_FORECAST
        ) t
        GROUP BY linha
        ORDER BY forecast DESC
    """)
    fc_por_linha = {}
    for r in cur.fetchall():
        fc_por_linha[r[0]] = {
            "forecast": float(r[1] or 0),
            "jan": float(r[2] or 0), "fev": float(r[3] or 0),
            "mar": float(r[4] or 0), "abr": float(r[5] or 0),
            "mai": float(r[6] or 0), "jun": float(r[7] or 0),
            "jul": float(r[8] or 0), "ago": float(r[9] or 0),
            "set_m": float(r[10] or 0), "out_m": float(r[11] or 0),
            "nov": float(r[12] or 0), "dez": float(r[13] or 0),
        }

    # Realizado por linha S&OP (DESCRICAO)
    _exec(cur, f"""
        SELECT linha,
               ISNULL(SUM(qtde),0) as real,
               ISNULL(SUM(jan),0) as jan, ISNULL(SUM(fev),0) as fev,
               ISNULL(SUM(mar),0) as mar, ISNULL(SUM(abr),0) as abr,
               ISNULL(SUM(mai),0) as mai, ISNULL(SUM(jun),0) as jun,
               ISNULL(SUM(jul),0) as jul, ISNULL(SUM(ago),0) as ago,
               ISNULL(SUM(set_m),0) as set_m, ISNULL(SUM(out_m),0) as out_m,
               ISNULL(SUM(nov),0) as nov, ISNULL(SUM(dez),0) as dez
        FROM (
            SELECT {_LINHA_CASE_DESCRICAO} as linha,
                   ISNULL(QUANTIDADE,0) as qtde,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=1  THEN QUANTIDADE END,0) as jan,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=2  THEN QUANTIDADE END,0) as fev,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=3  THEN QUANTIDADE END,0) as mar,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=4  THEN QUANTIDADE END,0) as abr,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=5  THEN QUANTIDADE END,0) as mai,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=6  THEN QUANTIDADE END,0) as jun,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=7  THEN QUANTIDADE END,0) as jul,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=8  THEN QUANTIDADE END,0) as ago,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=9  THEN QUANTIDADE END,0) as set_m,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=10 THEN QUANTIDADE END,0) as out_m,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=11 THEN QUANTIDADE END,0) as nov,
                   ISNULL(CASE WHEN MONTH(EMISSAO)=12 THEN QUANTIDADE END,0) as dez
            FROM FATURAMENTO
            WHERE YEAR(EMISSAO) = YEAR(GETDATE())
              AND TP NOT IN ('DEV','1DEV') AND TIPO NOT LIKE '%DEV%'
        ) t
        GROUP BY linha
    """)
    real_por_linha = {}
    for r in cur.fetchall():
        real_por_linha[r[0]] = {
            "real": float(r[1] or 0),
            "jan": float(r[2] or 0), "fev": float(r[3] or 0),
            "mar": float(r[4] or 0), "abr": float(r[5] or 0),
            "mai": float(r[6] or 0), "jun": float(r[7] or 0),
            "jul": float(r[8] or 0), "ago": float(r[9] or 0),
            "set_m": float(r[10] or 0), "out_m": float(r[11] or 0),
            "nov": float(r[12] or 0), "dez": float(r[13] or 0),
        }

    linhas = sorted(set(list(fc_por_linha.keys()) + list(real_por_linha.keys())))
    result = []
    for linha in linhas:
        if not linha:
            continue
        fc_data = fc_por_linha.get(linha, {})
        rl_data = real_por_linha.get(linha, {})
        fc = fc_data.get("forecast", 0.0)
        rl = rl_data.get("real", 0.0)
        result.append({
            "linha": linha,
            "forecast": fc,
            "real": rl,
            "atingimento": round((rl / fc * 100) if fc > 0 else 0, 1),
            "fc_mes": {m: fc_data.get(m, 0.0) for m in ["jan","fev","mar","abr","mai","jun","jul","ago","set_m","out_m","nov","dez"]},
            "real_mes": {m: rl_data.get(m, 0.0) for m in ["jan","fev","mar","abr","mai","jun","jul","ago","set_m","out_m","nov","dez"]},
        })
    return result


@router.get("/forecast/export")
def export_forecast(
    linha: str | None = Query(None),
    search: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    conds, params = [], []
    if linha and linha in _LINHA_COND:
        conds.append(_LINHA_COND[linha])
    if search:
        conds.append("(COD_PRD LIKE %s OR DESC_PRD LIKE %s)")
        p = f"%{search}%"; params += [p, p]
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    perc = _forecast_perc_col(bool(params))

    cur = conn.cursor()
    _exec(cur, f"""
        SELECT COD_PRD, DESC_PRD, FAM, LINHA, {perc} as PERC_ABC,
               JAN, FEV, MAR, ABR, MAI, JUN, JUL, AGO, [SET] as SET_M,
               OUT, NOV, DEZ, TOTAL
        FROM SKU_FORECAST {where}
        ORDER BY FAM, COD_PRD
    """, params)
    headers = ["Produto", "Descrição", "Família", "Linha", "% ABC",
               "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
               "Jul", "Ago", "Set", "Out", "Nov", "Dez", "Total"]
    xlsx = _build_xlsx("Forecast SKU", headers, [list(r) for r in cur.fetchall()])
    return _xlsx_response(xlsx, "forecast_sku.xlsx")


# ---------------------------------------------------------------------------
# ESTOQUE PA (ESTOQUES_PA)
# ---------------------------------------------------------------------------

def _estoque_where(empresa, linha, local, filial, familia, search):
    """ESTOQUES_PA: literais SQL escapados manualmente — valores vêm de listas controladas."""
    def _esc(v: str) -> str:
        return v.replace("'", "''")

    conds = []
    if empresa:
        conds.append(f"TB_EMPRESA = '{_esc(empresa)}'")
    if linha and linha in _LINHA_COND:
        conds.append(_LINHA_COND[linha])
    if local:
        conds.append(f"LOCAL = '{_esc(local)}'")
    if filial:                              # COD_FILIAL: '01'=Matriz, '06'=Atibaia
        conds.append(f"COD_FILIAL = '{_esc(filial)}'")
    if familia:
        conds.append(f"FAMILIA = '{_esc(familia)}'")
    if search:
        s = _esc(search)
        conds.append(f"(COD_PRD LIKE '%{s}%' OR DESC_PRD LIKE '%{s}%')")
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    return where, []


_FILIAL_LABEL = {"01": "Matriz (01)", "06": "Atibaia (06)"}

@router.get("/estoque")
def list_estoque(
    empresa: str | None = Query(None),
    linha: str | None = Query(None),
    local: str | None = Query(None),
    filial: str | None = Query(None),
    familia: str | None = Query(None),
    search: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, _params = _estoque_where(empresa, linha, local, filial, familia, search)
    cur = conn.cursor()

    # Itens — agrega por produto respeitando filtros ativos
    _exec(cur, f"""
        SELECT COD_PRD, DESC_PRD, ISNULL(FAMILIA,'') as FAMILIA,
               ISNULL(SUM(QTDE_ATU),0) as QTDE_ATU,
               ISNULL(SUM(QTDE_RESERVA),0) as QTDE_RESERVA,
               ISNULL(SUM(QTDE_EMP),0) as QTDE_EMP,
               ISNULL(SUM(QTDE_QACLASS),0) as QTDE_QACLASS,
               ISNULL(SUM(QTDE_DISP),0) as QTDE_DISP,
               '' as TB_EMPRESA, '' as EMPRESA, '' as LOCAL,
               0 as QTDE_EMPSAI, 0 as QTDE_EMPPRJ, 0 as QTDE_TNP, 0 as QTDE_EMPPRE
        FROM ESTOQUES_PA {where}
        GROUP BY COD_PRD, DESC_PRD, FAMILIA
        ORDER BY QTDE_DISP DESC
    """)
    raw_items = rows_to_dicts(cur)
    for item in raw_items:
        item["sk_estoque"] = item.get("cod_prd", "")
    items = raw_items

    # Por linha (respeitando filtros)
    _exec(cur, f"""
        SELECT linha,
               ISNULL(SUM(disp),0) as disp,
               ISNULL(SUM(reserva),0) as reserva,
               ISNULL(SUM(atu),0) as atu
        FROM (
            SELECT {_LINHA_CASE} as linha,
                   ISNULL(QTDE_DISP,0) as disp,
                   ISNULL(QTDE_RESERVA,0) as reserva,
                   ISNULL(QTDE_ATU,0) as atu
            FROM ESTOQUES_PA {where}
        ) t
        GROUP BY linha
        ORDER BY disp DESC
    """)
    por_linha = [
        {"linha": r[0], "disp": float(r[1] or 0),
         "reserva": float(r[2] or 0), "atu": float(r[3] or 0)}
        for r in cur.fetchall()
    ]

    # Breakdown por filial — sempre empresa 010, LOCAL='50'
    # COD_FILIAL '01'=Matriz, '06'=Atibaia; ignora filtros de local/empresa
    filial_extra = []
    if linha and linha in _LINHA_COND:
        filial_extra.append(_LINHA_COND[linha])
    if search:
        s = search.replace("'", "''")
        filial_extra.append(f"(COD_PRD LIKE '%{s}%' OR DESC_PRD LIKE '%{s}%')")
    filial_extra_sql = (" AND " + " AND ".join(filial_extra)) if filial_extra else ""

    _exec(cur, f"""
        SELECT COD_FILIAL,
               ISNULL(SUM(QTDE_ATU), 0)     as atu,
               ISNULL(SUM(QTDE_RESERVA), 0) as reserva,
               ISNULL(SUM(QTDE_DISP), 0)    as disp,
               ISNULL(SUM(QTDE_QACLASS), 0) as qaclass,
               COUNT(DISTINCT COD_PRD)       as produtos,
               SUM(CASE WHEN QTDE_DISP <= 0 THEN 1 ELSE 0 END) as zerados
        FROM ESTOQUES_PA
        WHERE TB_EMPRESA = '010' AND LOCAL = '50'
          AND COD_FILIAL IN ('01', '06')
          {filial_extra_sql}
        GROUP BY COD_FILIAL
        ORDER BY COD_FILIAL
    """)
    por_filial = [
        {
            "filial":   r[0],
            "label":    _FILIAL_LABEL.get(r[0], f"Filial {r[0]}"),
            "atu":      float(r[1] or 0),
            "reserva":  float(r[2] or 0),
            "disp":     float(r[3] or 0),
            "qaclass":  float(r[4] or 0),
            "produtos": int(r[5] or 0),
            "zerados":  int(r[6] or 0),
        }
        for r in cur.fetchall() if r[0]
    ]

    # KPIs agregados da seleção atual
    kpis = {
        "total_disp":    sum(r.get("qtde_disp", 0) or 0 for r in items),
        "total_atu":     sum(r.get("qtde_atu", 0) or 0 for r in items),
        "total_reserva": sum(r.get("qtde_reserva", 0) or 0 for r in items),
        "total_qaclass": sum(r.get("qtde_qaclass", 0) or 0 for r in items),
        "skus_total":    len(items),
        "skus_ativos":   sum(1 for r in items if (r.get("qtde_disp") or 0) > 0),
        "skus_zerados":  sum(1 for r in items if (r.get("qtde_disp") or 0) <= 0),
    }

    return {
        "items": items,
        "por_linha": por_linha,
        "por_filial": por_filial,
        "kpis": kpis,
    }


@router.get("/estoque/export")
def export_estoque(
    empresa: str | None = Query(None),
    linha: str | None = Query(None),
    local: str | None = Query(None),
    filial: str | None = Query(None),
    familia: str | None = Query(None),
    search: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, params = _estoque_where(empresa, linha, local, filial, familia, search)
    cur = conn.cursor()
    _exec(cur, f"""
        SELECT COD_PRD, DESC_PRD, ISNULL(FAMILIA,'') as FAMILIA, TB_EMPRESA,
               ISNULL(EMPRESA,'') as EMPRESA, LOCAL,
               ISNULL(QTDE_ATU,0), ISNULL(QTDE_RESERVA,0), ISNULL(QTDE_DISP,0),
               ISNULL(QTDE_EMP,0), ISNULL(QTDE_QACLASS,0)
        FROM ESTOQUES_PA {where}
        ORDER BY FAMILIA, COD_PRD
    """, params)
    headers = ["Produto", "Descrição", "Família", "Empresa (Cód)", "Empresa",
               "Local", "Qtde Atual", "Reserva", "Disponível", "Empréstimo", "QA/Bloq."]
    xlsx = _build_xlsx("Estoque PA", headers, [list(r) for r in cur.fetchall()])
    return _xlsx_response(xlsx, "estoque_pa.xlsx")


# ---------------------------------------------------------------------------
# FATURAMENTO
# ---------------------------------------------------------------------------

def _fat_where(linha, vendedor, uf, search, date_from, date_to, excluir_dev):
    conds, params = [], []
    if excluir_dev:
        conds.append("TP = 'VS'")  # apenas vendas normais (exclui DS=devolução e FS=bonificação)
    if linha and linha in _LINHA_COND_DESCRICAO:
        conds.append(_LINHA_COND_DESCRICAO[linha])
    if vendedor:
        conds.append("NOME_VEND = %s"); params.append(vendedor)
    if uf:
        conds.append("UF = %s"); params.append(uf)
    if search:
        conds.append("(COD_PRD LIKE %s OR DESCRICAO LIKE %s OR CLIENTE LIKE %s OR NF LIKE %s)")
        p = f"%{search}%"; params += [p, p, p, p]
    if date_from:
        conds.append("EMISSAO >= %s"); params.append(date_from)
    if date_to:
        conds.append("EMISSAO <= %s"); params.append(date_to)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    return where, params


@router.get("/faturamento")
def list_faturamento(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    linha: str | None = Query(None),
    vendedor: str | None = Query(None),
    uf: str | None = Query(None),
    search: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    excluir_dev: bool = Query(True),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, params = _fat_where(linha, vendedor, uf, search, date_from, date_to, excluir_dev)
    cur = conn.cursor()
    _exec(cur, f"SELECT COUNT(*) FROM FATURAMENTO {where}", params)
    total = cur.fetchone()[0]

    offset = (page - 1) * page_size
    _exec(cur, f"""
        SELECT SK_FAT, FILIAL, TP, FAMILIA, COD_PRD, DESCRICAO, NF, SERIE,
               CONVERT(varchar,EMISSAO,23) as EMISSAO,
               QUANTIDADE, CLIENTE, MUNICIPIO, UF,
               NOME_VEND, SUPERVISOR, VL_BRUTO, VL_LIQUIDO, PERC_COM, COM
        FROM FATURAMENTO {where}
        ORDER BY EMISSAO DESC, SK_FAT DESC
        OFFSET {offset} ROWS FETCH NEXT {page_size} ROWS ONLY
    """, params)
    items = rows_to_dicts(cur)
    return {
        "items": items, "total": total, "page": page,
        "page_size": page_size,
        "total_pages": max(1, math.ceil(total / page_size)),
    }


@router.get("/faturamento/charts")
def faturamento_charts(
    linha: str | None = Query(None),
    vendedor: str | None = Query(None),
    uf: str | None = Query(None),
    search: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    excluir_dev: bool = Query(True),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, params = _fat_where(linha, vendedor, uf, search, date_from, date_to, excluir_dev)
    cur = conn.cursor()

    _exec(cur, f"""
        SELECT FORMAT(EMISSAO,'yyyy-MM') as mes,
               ISNULL(SUM(VL_BRUTO),0) as bruto,
               ISNULL(SUM(VL_LIQUIDO),0) as liquido
        FROM FATURAMENTO {where}
        GROUP BY FORMAT(EMISSAO,'yyyy-MM') ORDER BY mes
    """, params)
    por_mes = [{"mes": r[0], "bruto": float(r[1] or 0), "liquido": float(r[2] or 0)}
               for r in cur.fetchall()]

    _exec(cur, f"""
        SELECT linha, ISNULL(SUM(valor),0) as valor
        FROM (
            SELECT {_LINHA_CASE_DESCRICAO} as linha,
                   ISNULL(VL_BRUTO,0) as valor
            FROM FATURAMENTO {where}
        ) t
        GROUP BY linha
        ORDER BY valor DESC
    """, params)
    por_linha = [{"linha": r[0], "valor": float(r[1] or 0)} for r in cur.fetchall()]

    _exec(cur, f"""
        SELECT TOP 10 NOME_VEND, ISNULL(SUM(VL_BRUTO),0) as valor
        FROM FATURAMENTO {where}
        GROUP BY NOME_VEND ORDER BY valor DESC
    """, params)
    por_vendedor = [{"vendedor": r[0], "valor": float(r[1] or 0)} for r in cur.fetchall()]

    return {"por_mes": por_mes, "por_linha": por_linha, "por_vendedor": por_vendedor}


@router.get("/faturamento/export")
def export_faturamento(
    linha: str | None = Query(None),
    vendedor: str | None = Query(None),
    uf: str | None = Query(None),
    search: str | None = Query(None),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    excluir_dev: bool = Query(True),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, params = _fat_where(linha, vendedor, uf, search, date_from, date_to, excluir_dev)
    cur = conn.cursor()
    _exec(cur, f"""
        SELECT TOP 50000
               CONVERT(varchar,EMISSAO,103), NF, SERIE, TP, FAMILIA,
               COD_PRD, DESCRICAO, QUANTIDADE, CLIENTE, MUNICIPIO, UF,
               NOME_VEND, SUPERVISOR, VL_BRUTO, VL_LIQUIDO, PERC_COM
        FROM FATURAMENTO {where}
        ORDER BY EMISSAO DESC
    """, params)
    headers = ["Emissão", "NF", "Série", "Tipo", "Família", "Produto", "Descrição",
               "Qtde", "Cliente", "Município", "UF",
               "Vendedor", "Supervisor", "Vl. Bruto", "Vl. Líquido", "% Comissão"]
    xlsx = _build_xlsx("Faturamento", headers, [list(r) for r in cur.fetchall()])
    return _xlsx_response(xlsx, "faturamento.xlsx")


@router.get("/faturamento/devolucoes")
def faturamento_devolucoes(
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    conds = ["TP = 'DS'"]  # DS = devolução (antigo filtro estava errado)
    params: list = []
    if date_from:
        conds.append("EMISSAO >= %s"); params.append(date_from)
    if date_to:
        conds.append("EMISSAO <= %s"); params.append(date_to)
    where = "WHERE " + " AND ".join(conds)

    cur = conn.cursor()

    # KPIs
    _exec(cur, f"""
        SELECT
            COUNT(*) as qtde_notas,
            ISNULL(SUM(QUANTIDADE),0) as qtde_itens,
            ISNULL(SUM(VL_LIQUIDO),0) as valor_total
        FROM FATURAMENTO {where}
    """, params)
    r = cur.fetchone()
    kpis = {
        "qtde_notas": int(r[0] or 0),
        "qtde_itens": float(r[1] or 0),
        "valor_total": float(r[2] or 0),
    }

    # Por período (mensal)
    _exec(cur, f"""
        SELECT FORMAT(EMISSAO,'yyyy-MM') as mes,
               ISNULL(SUM(VL_LIQUIDO),0) as valor,
               ISNULL(SUM(QUANTIDADE),0) as qtde
        FROM FATURAMENTO {where}
        GROUP BY FORMAT(EMISSAO,'yyyy-MM')
        ORDER BY mes
    """, params)
    por_periodo = [
        {"mes": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    # Por supervisor (área)
    _exec(cur, f"""
        SELECT ISNULL(SUPERVISOR,'(sem supervisor)') as supervisor,
               ISNULL(SUM(VL_LIQUIDO),0) as valor,
               ISNULL(SUM(QUANTIDADE),0) as qtde
        FROM FATURAMENTO {where}
        GROUP BY SUPERVISOR
        ORDER BY valor DESC
    """, params)
    por_supervisor = [
        {"supervisor": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    # Por UF
    _exec(cur, f"""
        SELECT ISNULL(UF,'??') as uf,
               ISNULL(SUM(VL_LIQUIDO),0) as valor
        FROM FATURAMENTO {where}
        GROUP BY UF
        ORDER BY valor DESC
    """, params)
    por_uf = [{"uf": r[0], "valor": float(r[1] or 0)} for r in cur.fetchall()]

    # Por linha S&OP
    _exec(cur, f"""
        SELECT linha, ISNULL(SUM(valor),0) as valor, ISNULL(SUM(qtde),0) as qtde
        FROM (
            SELECT {_LINHA_CASE_DESCRICAO} as linha,
                   ISNULL(VL_LIQUIDO,0) as valor,
                   ISNULL(QUANTIDADE,0) as qtde
            FROM FATURAMENTO {where}
        ) t
        GROUP BY linha
        ORDER BY valor DESC
    """, params)
    por_linha = [
        {"linha": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    return {
        "kpis": kpis,
        "por_periodo": por_periodo,
        "por_supervisor": por_supervisor,
        "por_uf": por_uf,
        "por_linha": por_linha,
    }


# ---------------------------------------------------------------------------
# DEVOLUÇÕES e BONIFICAÇÕES — visão dedicada
# TP='DS' → devolução  |  TP='FS' → bonificação (Bonific + Verba)
# VL_LIQUIDO nas DS é negativo → usar ABS()
# ---------------------------------------------------------------------------

def _devbonif_where(tp_prefix: str, date_from, date_to, linha, supervisor, uf):
    conds = [f"LEFT(TP,1) = '{tp_prefix}'"]
    params: list = []
    if date_from:
        conds.append("EMISSAO >= %s")
        params.append(date_from)
    if date_to:
        conds.append("EMISSAO <= %s")
        params.append(date_to)
    if linha and linha in _LINHA_COND_DESCRICAO:
        conds.append(_LINHA_COND_DESCRICAO[linha])
    if supervisor:
        conds.append("SUPERVISOR = %s")
        params.append(supervisor)
    if uf:
        conds.append("UF = %s")
        params.append(uf)
    return "WHERE " + " AND ".join(conds), params


@router.get("/devolucoes")
def get_devolucoes(
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    linha: str | None = Query(None),
    supervisor: str | None = Query(None),
    uf: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, params = _devbonif_where('D', date_from, date_to, linha, supervisor, uf)
    cur = conn.cursor()

    # KPIs
    _exec(cur, f"""
        SELECT COUNT(*) as cnt,
               ISNULL(SUM(ABS(QUANTIDADE)),0) as qtde,
               ISNULL(SUM(ABS(VL_LIQUIDO)),0) as valor
        FROM FATURAMENTO {where}
    """, params)
    r = cur.fetchone()
    kpis = {
        "total_nfs": int(r[0] or 0),
        "total_qtde": float(r[1] or 0),
        "valor_total": float(r[2] or 0),
    }

    # Por área responsável (AREDESC)
    _exec(cur, f"""
        SELECT ISNULL(NULLIF(LTRIM(RTRIM(AREDESC)),''),'(sem área)') as area,
               ISNULL(SUM(ABS(VL_LIQUIDO)),0) as valor,
               ISNULL(SUM(ABS(QUANTIDADE)),0) as qtde,
               COUNT(*) as cnt
        FROM FATURAMENTO {where}
        GROUP BY AREDESC
        ORDER BY valor DESC
    """, params)
    por_area = [
        {"area": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0), "cnt": int(r[3] or 0)}
        for r in cur.fetchall()
    ]

    # Por motivo (MOTDEST)
    _exec(cur, f"""
        SELECT ISNULL(NULLIF(LTRIM(RTRIM(MOTDEST)),''),'(sem motivo)') as motivo,
               ISNULL(SUM(ABS(VL_LIQUIDO)),0) as valor,
               ISNULL(SUM(ABS(QUANTIDADE)),0) as qtde,
               COUNT(*) as cnt
        FROM FATURAMENTO {where}
        GROUP BY MOTDEST
        ORDER BY valor DESC
    """, params)
    por_motivo = [
        {"motivo": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0), "cnt": int(r[3] or 0)}
        for r in cur.fetchall()
    ]

    # Por linha S&OP (DESCRICAO)
    _exec(cur, f"""
        SELECT linha, ISNULL(SUM(valor),0) as valor, ISNULL(SUM(qtde),0) as qtde
        FROM (
            SELECT {_LINHA_CASE_DESCRICAO} as linha,
                   ISNULL(ABS(VL_LIQUIDO),0) as valor,
                   ISNULL(ABS(QUANTIDADE),0) as qtde
            FROM FATURAMENTO {where}
        ) t
        GROUP BY linha
        ORDER BY valor DESC
    """, params)
    por_linha = [
        {"linha": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    # Por supervisor
    _exec(cur, f"""
        SELECT ISNULL(SUPERVISOR,'(sem supervisor)') as supervisor,
               ISNULL(SUM(ABS(VL_LIQUIDO)),0) as valor,
               ISNULL(SUM(ABS(QUANTIDADE)),0) as qtde,
               COUNT(*) as cnt
        FROM FATURAMENTO {where}
        GROUP BY SUPERVISOR
        ORDER BY valor DESC
    """, params)
    por_supervisor = [
        {"supervisor": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0), "cnt": int(r[3] or 0)}
        for r in cur.fetchall()
    ]

    # Por UF
    _exec(cur, f"""
        SELECT ISNULL(UF,'??') as uf,
               ISNULL(SUM(ABS(VL_LIQUIDO)),0) as valor,
               COUNT(*) as cnt
        FROM FATURAMENTO {where}
        GROUP BY UF
        ORDER BY valor DESC
    """, params)
    por_uf = [{"uf": r[0], "valor": float(r[1] or 0), "cnt": int(r[2] or 0)} for r in cur.fetchall()]

    # Por período (mensal)
    _exec(cur, f"""
        SELECT FORMAT(EMISSAO,'yyyy-MM') as mes,
               ISNULL(SUM(ABS(VL_LIQUIDO)),0) as valor,
               ISNULL(SUM(ABS(QUANTIDADE)),0) as qtde,
               COUNT(*) as cnt
        FROM FATURAMENTO {where}
        GROUP BY FORMAT(EMISSAO,'yyyy-MM')
        ORDER BY mes
    """, params)
    por_periodo = [
        {"mes": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0), "cnt": int(r[3] or 0)}
        for r in cur.fetchall()
    ]

    return {
        "kpis": kpis,
        "por_area": por_area,
        "por_motivo": por_motivo,
        "por_linha": por_linha,
        "por_supervisor": por_supervisor,
        "por_uf": por_uf,
        "por_periodo": por_periodo,
    }


@router.get("/bonificacoes")
def get_bonificacoes(
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    linha: str | None = Query(None),
    supervisor: str | None = Query(None),
    uf: str | None = Query(None),
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    where, params = _devbonif_where('F', date_from, date_to, linha, supervisor, uf)
    cur = conn.cursor()

    # KPIs + breakdown por subtipo (TIPO: 'Bonific' vs 'Verba')
    _exec(cur, f"""
        SELECT COUNT(*) as cnt,
               ISNULL(SUM(QUANTIDADE),0) as qtde,
               ISNULL(SUM(VL_LIQUIDO),0) as valor,
               ISNULL(SUM(CASE WHEN TIPO='Bonific' THEN VL_LIQUIDO END),0) as vl_bonific,
               ISNULL(SUM(CASE WHEN TIPO='Verba'   THEN VL_LIQUIDO END),0) as vl_verba
        FROM FATURAMENTO {where}
    """, params)
    r = cur.fetchone()
    kpis = {
        "total_nfs": int(r[0] or 0),
        "total_qtde": float(r[1] or 0),
        "valor_total": float(r[2] or 0),
        "vl_bonific": float(r[3] or 0),
        "vl_verba": float(r[4] or 0),
    }

    # Por linha S&OP
    _exec(cur, f"""
        SELECT linha, ISNULL(SUM(valor),0) as valor, ISNULL(SUM(qtde),0) as qtde
        FROM (
            SELECT {_LINHA_CASE_DESCRICAO} as linha,
                   ISNULL(VL_LIQUIDO,0) as valor,
                   ISNULL(QUANTIDADE,0) as qtde
            FROM FATURAMENTO {where}
        ) t
        GROUP BY linha
        ORDER BY valor DESC
    """, params)
    por_linha = [
        {"linha": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    # Por supervisor
    _exec(cur, f"""
        SELECT ISNULL(SUPERVISOR,'(sem supervisor)') as supervisor,
               ISNULL(SUM(VL_LIQUIDO),0) as valor,
               ISNULL(SUM(QUANTIDADE),0) as qtde,
               COUNT(*) as cnt
        FROM FATURAMENTO {where}
        GROUP BY SUPERVISOR
        ORDER BY valor DESC
    """, params)
    por_supervisor = [
        {"supervisor": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0), "cnt": int(r[3] or 0)}
        for r in cur.fetchall()
    ]

    # Por subtipo (Bonific vs Verba)
    _exec(cur, f"""
        SELECT ISNULL(TIPO,'?') as subtipo,
               ISNULL(SUM(VL_LIQUIDO),0) as valor,
               ISNULL(SUM(QUANTIDADE),0) as qtde,
               COUNT(*) as cnt
        FROM FATURAMENTO {where}
        GROUP BY TIPO
        ORDER BY valor DESC
    """, params)
    por_subtipo = [
        {"subtipo": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0), "cnt": int(r[3] or 0)}
        for r in cur.fetchall()
    ]

    # Top clientes
    _exec(cur, f"""
        SELECT TOP 20 ISNULL(CLIENTE,'(sem cliente)') as cliente,
               ISNULL(SUM(VL_LIQUIDO),0) as valor,
               ISNULL(SUM(QUANTIDADE),0) as qtde
        FROM FATURAMENTO {where}
        GROUP BY CLIENTE
        ORDER BY valor DESC
    """, params)
    por_cliente = [
        {"cliente": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0)}
        for r in cur.fetchall()
    ]

    # Por UF
    _exec(cur, f"""
        SELECT ISNULL(UF,'??') as uf,
               ISNULL(SUM(VL_LIQUIDO),0) as valor,
               COUNT(*) as cnt
        FROM FATURAMENTO {where}
        GROUP BY UF
        ORDER BY valor DESC
    """, params)
    por_uf = [{"uf": r[0], "valor": float(r[1] or 0), "cnt": int(r[2] or 0)} for r in cur.fetchall()]

    # Por período
    _exec(cur, f"""
        SELECT FORMAT(EMISSAO,'yyyy-MM') as mes,
               ISNULL(SUM(VL_LIQUIDO),0) as valor,
               ISNULL(SUM(QUANTIDADE),0) as qtde,
               COUNT(*) as cnt
        FROM FATURAMENTO {where}
        GROUP BY FORMAT(EMISSAO,'yyyy-MM')
        ORDER BY mes
    """, params)
    por_periodo = [
        {"mes": r[0], "valor": float(r[1] or 0), "qtde": float(r[2] or 0), "cnt": int(r[3] or 0)}
        for r in cur.fetchall()
    ]

    return {
        "kpis": kpis,
        "por_subtipo": por_subtipo,
        "por_linha": por_linha,
        "por_supervisor": por_supervisor,
        "por_uf": por_uf,
        "por_periodo": por_periodo,
        "por_cliente": por_cliente,
    }


# ---------------------------------------------------------------------------
# PMP vs REAL — Programado (ZPM010) x Realizado (MOV_PRODUCAO) por SKU
# ---------------------------------------------------------------------------

@router.get("/pmp-real/meses")
def pmp_real_meses(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Retorna meses e linhas disponíveis no ZPM010."""
    meses_rows = db.execute(text(
        "SELECT DISTINCT ZPM_MESREF FROM ZPM010"
        " WHERE D_E_L_E_T_ <> '*' AND LTRIM(RTRIM(ZPM_MESREF)) <> ''"
        " ORDER BY ZPM_MESREF DESC"
    )).fetchall()
    linhas_rows = db.execute(text(
        "SELECT DISTINCT ZPM_LINHA FROM ZPM010"
        " WHERE D_E_L_E_T_ <> '*' AND LTRIM(RTRIM(ZPM_LINHA)) <> ''"
        " ORDER BY ZPM_LINHA"
    )).fetchall()
    return {
        "meses": [r[0].strip() for r in meses_rows if r[0]],
        "linhas": [r[0].strip() for r in linhas_rows if r[0]],
    }


@router.get("/pmp-real")
def pmp_real(
    mesref: str | None = Query(None),
    linha: str | None = Query(None),
    search: str | None = Query(None),
    conn=Depends(get_bi_conn),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if not mesref:
        now = datetime.date.today()
        mesref = f"{now.year}{now.month:02d}"

    ano = int(mesref[:4])
    mes = int(mesref[4:6])

    # 1. PMP from ZPM010 (main DB via SQLAlchemy)
    day_cols = ", ".join([f"ZPM_D{i:02d}" for i in range(1, 32)])
    sql_pmp = text(
        f"SELECT ZPM_PROD, ZPM_DESC, ZPM_LINHA, ZPM_PMPMES, {day_cols}"
        " FROM ZPM010"
        " WHERE ZPM_MESREF = :mesref AND D_E_L_E_T_ <> '*'"
        " ORDER BY ZPM_LINHA, ZPM_PROD"
    )
    pmp_rows = db.execute(sql_pmp, {"mesref": mesref}).fetchall()

    # 2. Produção real from BI DB (MOV_PRODUCAO)
    cur = conn.cursor()
    _exec(cur, """
        SELECT COD_PRD, DAY(EMISSAO) as dia, ISNULL(SUM(QTDE),0) as qtde
        FROM MOV_PRODUCAO
        WHERE TIPO = 'PA'
          AND YEAR(EMISSAO) = %s AND MONTH(EMISSAO) = %s
        GROUP BY COD_PRD, DAY(EMISSAO)
    """, [ano, mes])
    real_map: dict[str, dict[int, float]] = {}
    for row in cur.fetchall():
        cod = (row[0] or "").strip()
        if cod:
            real_map.setdefault(cod, {})[int(row[1])] = float(row[2] or 0)

    # 3. Combine
    items = []
    for row in pmp_rows:
        cod_prd = (row[0] or "").strip()
        desc    = (row[1] or "").strip()
        linha_prd = (row[2] or "").strip()
        pmpmes  = float(row[3] or 0)
        dias_prog = [float(row[4 + i] or 0) for i in range(31)]

        if linha and linha_prd != linha:
            continue
        if search:
            s = search.lower()
            if s not in cod_prd.lower() and s not in desc.lower():
                continue

        real_dias = real_map.get(cod_prd, {})
        dias_real = [real_dias.get(i + 1, 0.0) for i in range(31)]

        programado = sum(dias_prog)
        realizado  = sum(dias_real)

        items.append({
            "cod_prd":    cod_prd,
            "desc":       desc,
            "linha":      linha_prd,
            "pmpmes":     pmpmes,
            "programado": programado,
            "realizado":  realizado,
            "atingimento": round(realizado / programado * 100, 1) if programado > 0 else 0.0,
            "prog": {f"d{i+1:02d}": dias_prog[i] for i in range(31)},
            "real": {f"d{i+1:02d}": dias_real[i] for i in range(31)},
        })

    total_prog = sum(x["programado"] for x in items)
    total_real = sum(x["realizado"] for x in items)

    return {
        "mesref": mesref,
        "kpis": {
            "programado":  total_prog,
            "realizado":   total_real,
            "atingimento": round(total_real / total_prog * 100, 1) if total_prog > 0 else 0.0,
        },
        "items": items,
    }


# ---------------------------------------------------------------------------
# PROG. EMBARQUE — Ruptura por produto (próximos N dias)
# ---------------------------------------------------------------------------

@router.get("/embarque")
def get_embarque(
    dias: int = Query(7, ge=1, le=14),
    linha: str | None = Query(None),
    search: str | None = Query(None),
    conn=Depends(get_bi_conn),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    today = datetime.date.today()
    datas = [today + datetime.timedelta(days=i) for i in range(dias)]
    datas_str = [d.strftime('%Y-%m-%d') for d in datas]
    date_to = datas[-1]

    cur = conn.cursor()

    # Condições de filtro
    extra_conds = []
    extra_params: list = []
    if linha and linha in _LINHA_COND:
        extra_conds.append(_LINHA_COND[linha])
    if search:
        extra_conds.append("(COD_PRD LIKE %s OR DESC_PRD LIKE %s)")
        p = f"%{search}%"
        extra_params += [p, p]

    extra_sql = (" AND " + " AND ".join(extra_conds)) if extra_conds else ""

    # 1. Demanda por produto por dia de carregamento
    _exec(cur, f"""
        SELECT COD_PRD, DESC_PRD,
               CAST(DT_CARREGAMENTO AS DATE) as dt,
               SUM(QTDE) as demanda,
               COUNT(DISTINCT NUM_PV) as qtd_pvs
        FROM PV_ABERTOS
        WHERE DT_CARREGAMENTO >= %s
          AND DT_CARREGAMENTO <= %s
          AND DT_CARREGAMENTO > '1990-01-01'
          {extra_sql}
        GROUP BY COD_PRD, DESC_PRD, CAST(DT_CARREGAMENTO AS DATE)
    """, [today.strftime('%Y-%m-%d'), date_to.strftime('%Y-%m-%d')] + extra_params)
    demanda_raw = rows_to_dicts(cur)

    if not demanda_raw:
        return {
            "datas": datas_str,
            "items": [],
            "kpis": {"total_prds": 0, "em_ruptura": 0, "em_alerta": 0, "demanda_total": 0},
        }

    # Coleta produtos e monta matriz de demanda (normaliza espaços do código)
    produtos_info: dict[str, str] = {}
    demanda_matrix: dict[str, dict[str, float]] = {}
    for r in demanda_raw:
        cod = (r['cod_prd'] or "").strip()
        if not cod:
            continue
        produtos_info.setdefault(cod, (r['desc_prd'] or "").strip())
        demanda_matrix.setdefault(cod, {})[str(r['dt'])] = float(r['demanda'] or 0)

    produtos = list(produtos_info.keys())

    # 2. Estoque disponível (empresa 010)
    ph = ','.join(['%s'] * len(produtos))
    _exec(cur, f"""
        SELECT COD_PRD, ISNULL(SUM(QTDE_DISP), 0) as disp
        FROM ESTOQUES_PA
        WHERE TB_EMPRESA = '010'
          AND COD_PRD IN ({ph})
        GROUP BY COD_PRD
    """, produtos)
    estoque_map: dict[str, float] = {(r[0] or "").strip(): float(r[1]) for r in cur.fetchall()}

    # 3. PMP do ZPM010 (SQLAlchemy — banco principal)
    meses_needed = set(d.strftime('%Y%m') for d in datas)
    pmp_matrix: dict[str, dict[str, float]] = {}

    for mesref_zpm in sorted(meses_needed):
        yr = int(mesref_zpm[:4])
        mo = int(mesref_zpm[4:])
        day_cols = ", ".join([f"ZPM_D{i:02d}" for i in range(1, 32)])
        prod_in = "', '".join(p.replace("'", "''") for p in produtos)
        rows_pmp = db.execute(text(f"""
            SELECT LTRIM(RTRIM(ZPM_PROD)) as ZPM_PROD, {day_cols}
            FROM ZPM010
            WHERE ZPM_MESREF = :mesref AND D_E_L_E_T_ <> '*'
              AND LTRIM(RTRIM(ZPM_PROD)) IN ('{prod_in}')
        """), {"mesref": mesref_zpm}).fetchall()

        for row in rows_pmp:
            cod = (row[0] or "").strip()
            if not cod:
                continue
            pmp_matrix.setdefault(cod, {})
            for day_idx in range(1, 32):
                try:
                    day_date = datetime.date(yr, mo, day_idx)
                except ValueError:
                    break
                val = row[day_idx]  # ZPM_D01=index 1, ZPM_D02=index 2, ...
                pmp_matrix[cod][day_date.strftime('%Y-%m-%d')] = float(val or 0)

    # 4. Linha por produto (a partir de PV_ABERTOS)
    _exec(cur, f"""
        SELECT DISTINCT COD_PRD, {_LINHA_CASE} as linha
        FROM PV_ABERTOS
        WHERE COD_PRD IN ({ph})
    """, produtos)
    linha_map: dict[str, str] = {(r[0] or "").strip(): r[1] for r in cur.fetchall()}

    # 5. Calcular saldo cumulativo e status
    STATUS_ORDER = {"RUPTURA": 0, "ALERTA": 1, "OK": 2}
    items = []

    for cod in sorted(produtos):
        estoque_disp = estoque_map.get(cod, 0.0)
        dem_by_day = demanda_matrix.get(cod, {})
        pmp_by_day = pmp_matrix.get(cod, {})
        demanda_total = sum(dem_by_day.values())
        media_diaria = demanda_total / len(datas_str) if datas_str else 0

        dias_data = []
        saldo = estoque_disp
        tem_ruptura = False
        tem_alerta = False

        for dt_str in datas_str:
            dem = dem_by_day.get(dt_str, 0.0)
            pmp = pmp_by_day.get(dt_str, 0.0)
            saldo = saldo - dem + pmp

            dias_data.append({
                "data": dt_str,
                "demanda": dem,
                "pmp": pmp,
                "saldo": round(saldo),
            })

            if saldo < 0:
                tem_ruptura = True
            elif media_diaria > 0 and 0 <= saldo < media_diaria:
                tem_alerta = True

        status = "RUPTURA" if tem_ruptura else ("ALERTA" if tem_alerta else "OK")

        items.append({
            "cod_prd": cod,
            "desc_prd": produtos_info[cod],
            "linha": linha_map.get(cod, "Outros"),
            "estoque_disp": estoque_disp,
            "demanda_total": demanda_total,
            "dias": dias_data,
            "status": status,
        })

    items.sort(key=lambda x: (STATUS_ORDER[x["status"]], -x["demanda_total"]))

    kpis = {
        "total_prds": len(items),
        "em_ruptura": sum(1 for i in items if i["status"] == "RUPTURA"),
        "em_alerta": sum(1 for i in items if i["status"] == "ALERTA"),
        "demanda_total": round(sum(i["demanda_total"] for i in items)),
    }

    return {"datas": datas_str, "items": items, "kpis": kpis}


# ---------------------------------------------------------------------------
# FILTROS — opções únicas para os selects do frontend
# ---------------------------------------------------------------------------

@router.get("/filters")
def get_filters(
    conn=Depends(get_bi_conn),
    current_user: models.User = Depends(get_current_user),
):
    cur = conn.cursor()

    def _distinct(sql):
        try:
            _exec(cur, sql)
            return [r[0] for r in cur.fetchall() if r[0]]
        except Exception:
            return []

    def _empresas():
        try:
            _exec(cur,
                  "SELECT DISTINCT TB_EMPRESA, EMPRESA "
                  "FROM ESTOQUES_PA "
                  "WHERE TB_EMPRESA IS NOT NULL "
                  "ORDER BY TB_EMPRESA")
        except Exception:
            return []
        return [{"cod": r[0], "nome": r[1]} for r in cur.fetchall() if r[0]]

    # Descrição de família: cruza FAM de PV_ABERTOS com FAMILIA de FATURAMENTO via COD_PRD
    try:
        _exec(cur, """
            SELECT p.FAM, MIN(f.FAMILIA) as label
            FROM (
                SELECT DISTINCT FAM, COD_PRD FROM PV_ABERTOS
                WHERE FAM IS NOT NULL AND FAM <> ''
            ) p
            INNER JOIN (
                SELECT DISTINCT COD_PRD, FAMILIA FROM FATURAMENTO
                WHERE FAMILIA IS NOT NULL AND FAMILIA <> ''
            ) f ON p.COD_PRD = f.COD_PRD
            GROUP BY p.FAM
            ORDER BY p.FAM
        """)
        rows = cur.fetchall()
        if rows:
            familias_pv = [{"code": r[0], "label": r[1] or r[0]} for r in rows if r[0]]
        else:
            raise ValueError("no rows")
    except Exception:
        # fallback: sem descrição, usa o código como label
        _exec(cur, "SELECT DISTINCT FAM FROM PV_ABERTOS WHERE FAM IS NOT NULL AND FAM<>'' ORDER BY FAM")
        familias_pv = [{"code": r[0], "label": r[0]} for r in cur.fetchall() if r[0]]

    return {
        "familias_pv": familias_pv,
        "linhas_pv": LINHAS_PV,
        "formatos_pv": FORMATOS_PV,
        "vendedores_pv": _distinct(
            "SELECT DISTINCT NOME_VEND FROM PV_ABERTOS "
            "WHERE NOME_VEND IS NOT NULL AND NOME_VEND<>'' ORDER BY NOME_VEND"
        ),
        "tipos_producao": _distinct(
            "SELECT DISTINCT TIPO FROM MOV_PRODUCAO "
            "WHERE TIPO IS NOT NULL ORDER BY TIPO"
        ),
        "familias_forecast": _distinct(
            "SELECT DISTINCT FAM FROM SKU_FORECAST "
            "WHERE FAM IS NOT NULL ORDER BY FAM"
        ),
        "familias_estoque": _distinct(
            "SELECT DISTINCT FAMILIA FROM ESTOQUES_PA "
            "WHERE FAMILIA IS NOT NULL AND FAMILIA<>'' ORDER BY FAMILIA"
        ),
        "locais_estoque": _distinct(
            "SELECT DISTINCT LOCAL FROM ESTOQUES_PA "
            "WHERE LOCAL IS NOT NULL AND LOCAL<>'' ORDER BY LOCAL"
        ),
        "empresas_estoque": _empresas(),
        "familias_faturamento": _distinct(
            "SELECT DISTINCT FAMILIA FROM FATURAMENTO "
            "WHERE FAMILIA IS NOT NULL AND FAMILIA<>'' ORDER BY FAMILIA"
        ),
        "vendedores_faturamento": _distinct(
            "SELECT DISTINCT NOME_VEND FROM FATURAMENTO "
            "WHERE NOME_VEND IS NOT NULL AND NOME_VEND<>'' ORDER BY NOME_VEND"
        ),
        "ufs": _distinct(
            "SELECT DISTINCT UF FROM FATURAMENTO "
            "WHERE UF IS NOT NULL AND UF<>'' ORDER BY UF"
        ),
    }
